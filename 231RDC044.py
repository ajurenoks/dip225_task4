import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

import time
from openpyxl import Workbook, load_workbook 

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

name=[]
# program read information from people.csv file and put all data in name list.
with open("people.csv", "r") as file:
    next(file)
    for line in file:
        row=line.rstrip().split(",") 
        name.append(row)

for i in range (0, len(name)):
        name[i]=name[i][2]+" "+name[i][3]
        

url = "https://emn178.github.io/online-tools/crc32.html"
driver.get(url)
codedName = [] 

for i in range(0, len(name)):
    find = driver.find_element(By.ID, "input")
    temp=name[i]
    find.send_keys(temp)
    find = driver.find_element(By.ID, "output")
    temp = find.get_attribute("value")
    codedName.append(temp)
    driver.find_element(By.ID, "input").clear()

wb = load_workbook ('salary.xlsx')
ws= wb.active
max_row = ws.max_row


for i in range(0, len(codedName)):
    totalSalary = 0
    for j in range(2, max_row+1):
        if(codedName[i]==ws['A'+str(j)].value):
           totalSalary=totalSalary+int(ws['B'+str(j)].value)
    name[i]=name[i]+" "+ str(totalSalary)

input=input("input name and surname of worker to know the salary: ")
for i in range(0, len(name)):
    if input in name[i]:
        print(name[i])

